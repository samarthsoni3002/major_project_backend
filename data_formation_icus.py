from firebase_admin import credentials, db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import firebase_admin
from firebase_admin import credentials, db

def fetch_icu_data():
   
 
    if not firebase_admin._apps:
        cred = credentials.Certificate('major_project_backend/hand_wash_json_file.json')
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://handhygiene-jaypeehealthcare-default-rtdb.firebaseio.com/'
        })

    # Configuration
    target_icus = ["ACTVs", "MICU", "NICU", "PICU", "TICU"]
    target_roles = ["Doctors", "GDA", "Nurse", "Para"]

    # Fetch data from Firebase
    root_ref = db.reference('/')
    data = root_ref.child('Count').child('1').get()

    # Create Excel workbook
    output_file = "major_project_backend/HandHygieneData_with_Styled_Graphs.xlsx"
    workbook = Workbook()

    # Styling configurations
    main_heading_font = Font(bold=True, color='FFFFFF', size=14)
    role_heading_font = Font(bold=True, color='FFFFFF')
    data_font = Font(size=12)
    main_heading_fill = PatternFill(start_color='4BACC6', end_color='4BACC6', fill_type='solid')
    role_heading_fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    red_font = Font(color='FF0000')

    # Helper function to auto-adjust column width
    def auto_adjust_column_width(sheet):
        for col in sheet.columns:
            values = [len(str(cell.value)) for cell in col if cell.value is not None]
            if values:
                max_length = max(values)
            else:
                max_length = 0
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # Iterate over ICUs and create sheets for each
    for icu in target_icus:
        icu_data = data.get(icu, {})
        if not icu_data:
            continue

        # Create a new sheet for each ICU
        sheet = workbook.create_sheet(title=icu)

        # Add ICU name as the main heading
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        sheet.cell(row=1, column=1, value=f"Hand Hygiene Compliance Data - {icu}").font = main_heading_font
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=1, column=1).fill = main_heading_fill

        # Leave space before adding the data
        current_row = 4

        for role in target_roles:
            role_data = icu_data.get(role, {})
            if not role_data:
                continue

            # Write role data header
            sheet.cell(row=current_row, column=1, value=f"{role} Data")
            sheet.cell(row=current_row, column=1).font = role_heading_font
            sheet.cell(row=current_row, column=1).fill = role_heading_fill
            sheet.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            current_row += 1

            # Write column headers
            sheet.cell(row=current_row, column=1, value="Class").font = Font(bold=True)
            col = 2
            metrics = next(iter(role_data.values())).keys()
            for metric in metrics:
                sheet.cell(row=current_row, column=col, value=metric).font = Font(bold=True)
                col += 1
            current_row += 1

            # Write class data rows
            for class_name, metrics in role_data.items():
                sheet.cell(row=current_row, column=1, value=class_name)
                col = 2
                for value in metrics.values():
                    cell = sheet.cell(row=current_row, column=col, value=value)
                    if value == 0:
                        cell.font = red_font  # Make the font color red if value is 0
                    cell.border = thin_border
                    col += 1
                current_row += 1

        # Auto-adjust column width for better readability
        auto_adjust_column_width(sheet)

        # Set specific column widths for G, H, I, J, K
        for col in range(6, 12):  # Columns G (7) to K (11)
            sheet.column_dimensions[get_column_letter(col)].width = 10

        # Summarize data for graph creation
        summary = {metric: 0 for metric in next(iter(next(iter(icu_data.values())).values())).keys()}
        for role, role_data in icu_data.items():
            for class_data in role_data.values():
                for metric, value in class_data.items():
                    if metric not in ["ChangedandHandHygieneDone", "NotUsed", "RemovedandHandHygieneDone", "Wornthroughouttheprocedure"]:
                        summary[metric] += value

        # Write summary data to columns G, H, I, J, K with green background
        graph_data_start_row = 2
        graph_data_col_start = 7
        for i, (metric, count) in enumerate(summary.items(), start=1):
            metric_cell = sheet.cell(row=graph_data_start_row + i, column=graph_data_col_start, value=metric)
            metric_cell.fill = green_fill
            metric_cell.border = thin_border

            count_cell = sheet.cell(row=graph_data_start_row + i, column=graph_data_col_start + 1, value=count)
            count_cell.fill = green_fill
            count_cell.border = thin_border

        # Generate the bar chart
        bar_chart = BarChart()
        bar_chart.title = f"{icu} Role-Wise Compliance"
        bar_chart.style = 12
        bar_chart.x_axis.title = "Metrics"
        bar_chart.y_axis.title = "Counts"
        bar_chart.width = 15
        bar_chart.height = 8

        # Reference data for the chart
        data_ref = Reference(sheet, min_col=graph_data_col_start + 1, min_row=graph_data_start_row + 1,
                            max_row=graph_data_start_row + len(summary))
        categories_ref = Reference(sheet, min_col=graph_data_col_start, min_row=graph_data_start_row + 1,
                                max_row=graph_data_start_row + len(summary))

        bar_chart.add_data(data_ref, titles_from_data=False)
        bar_chart.set_categories(categories_ref)
        
        # Add data labels
        bar_chart.dLbls = DataLabelList()
        bar_chart.dLbls.showVal = True

        # Adjust the chart position (from column K onward)
        graph_row_start = graph_data_start_row + len(summary) + 2
        graph_col_start = 11  # Column K (index 11)
        sheet.add_chart(bar_chart, f"{get_column_letter(graph_col_start)}{graph_row_start}")

    # Save the workbook
    workbook.save(output_file)
    print(f"Excel file created with styled graphs and green-highlighted data: {output_file}")
