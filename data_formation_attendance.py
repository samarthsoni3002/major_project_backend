import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import firebase_admin
from firebase_admin import credentials, db

def fetch_attendance_data():
   
    # Function to fetch data from Firebase
    def fetch_attendance_data():
        ref = db.reference('Attendance')
        attendance_data = ref.get()
        return attendance_data if attendance_data else {}

    # Apply formatting to a section header
    def format_section_header(worksheet, row, title):
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = worksheet.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply formatting to a table header
    def format_table_header(worksheet, start_row):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[start_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Add a bar chart for attendance
    def add_attendance_chart(worksheet, start_row, end_row, category):
        chart = BarChart()
        chart.title = f"{category} Attendance Analysis"
        chart.style = 2
        chart.x_axis.title = "ICU Types"
        chart.y_axis.title = "Count"

        data = Reference(worksheet, min_col=2, min_row=start_row, max_row=end_row, max_col=2)
        categories = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 15
        worksheet.add_chart(chart, f"F{start_row}")

    # Process and write data to an existing Excel file
    def process_and_write_to_excel(input_excel_path, output_excel_path):
        try:
            # Load the existing Excel workbook
            workbook = load_workbook(input_excel_path)
        except Exception as e:
            print(f"Error loading the Excel file: {e}")
            return

        # Fetch attendance data from Firebase
        attendance_data = fetch_attendance_data()
        if not attendance_data:
            print("No attendance data found in Firebase.")
            return
        
        # Categories to process
        categories = ["Doctors", "Nurse", "Para", "GDA"]
        
        for category in categories:
            if category in attendance_data:
                category_data = attendance_data[category]
                
                # Create a new sheet for the category
                sheet_name = f"{category}_attendance"
                if sheet_name in workbook.sheetnames:
                    print(f"Sheet '{sheet_name}' already exists, skipping.")
                    continue
                
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # Add the sheet name as the heading
                format_section_header(worksheet, 1, f"{category} Data")
                
                current_row = 3  # Start after heading and spacing
                
                # Process each name in the category
                for name, details in category_data.items():
                    # Add name as a sub-heading
                    worksheet.cell(row=current_row, column=1, value=f"{name}")
                    worksheet.cell(row=current_row, column=1).font = Font(bold=True, italic=True, size=12)
                    current_row += 1  # Leave one row below the name heading
                    
                    # Add table headers
                    worksheet.cell(row=current_row, column=1, value="ICU Type")
                    worksheet.cell(row=current_row, column=2, value="Attendance")
                    format_table_header(worksheet, current_row)  # Apply styling to headers
                    
                    current_row += 1  # Move to the next row for data
                    
                    # Add data rows
                    for icu_type, attendance in details.items():
                        worksheet.cell(row=current_row, column=1, value=icu_type)
                        if isinstance(attendance, dict):
                            worksheet.cell(row=current_row, column=2, value=attendance.get("HandRub", 0))
                        else:
                            worksheet.cell(row=current_row, column=2, value=attendance)
                        current_row += 1
                    
                    # Add space after each table
                    current_row += 2
                
                # Add a chart for the current category
                chart_start_row = 6  # Adjust based on your layout
                add_attendance_chart(worksheet, chart_start_row, current_row - 3, category)
        
        # Save the updated workbook
        workbook.save(output_excel_path)
        print(f"Updated Excel file saved as: {output_excel_path}")

    # Usage
    input_excel_path = "Backend_majorproject/HandHygieneData_with_Styled_Graphs.xlsx"  
    output_excel_path = "Backend_majorproject/HandHygiene_Compliance_Report.xlsx" 
    process_and_write_to_excel(input_excel_path, output_excel_path)
    return output_excel_path
